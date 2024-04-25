from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


def main(result):
    # Cargar el archivo de Excel
    workbook = load_workbook(filename='Pruebas.xlsx')

    # Seleccionar la hoja de Excel
    sheet = workbook.active

    # Obtener la letra de la columna para la siguiente celda disponible
    next_column = get_column_letter(sheet.max_column + 1)

    # Obtener la fila para el resultado actual
    row = sheet.max_row + 1

    # Escribir el resultado en la celda correspondiente
    sheet[next_column + str(row)] = result

    # Guardar los cambios en el archivo Excel
    workbook.save(filename='Pruebas.xlsx')

    # Cerrar el archivo de Excel
    workbook.close()


# Llamada a la función main() con el resultado de la prueba
if __name__ == "__main__":
    # Suponiendo que `result` contiene el resultado de la prueba
    result = "Prueba exitosa"  # o "Prueba fallida", dependiendo del resultado real
    main(result)
